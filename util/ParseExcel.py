import openpyxl
from openpyxl.styles import Border,Side,Font
import time

class PaseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        return sheet.max_row

    def getColsNumber(self,sheet):
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        return sheet.min_row

    def getStartColNumber(self,sheet):
        return sheet.min_colum

    def getRow(self,sheet,rowNo):
        try:
            return sheet.rows[rowNo - 1]  #下标从1开始row[1]表示第一行
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        try:
            return sheet.columns[colNo - 1]  # 下标从1开始column[1]表示第一列
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        if coordinate!=None:
            try:
                return sheet.cell(coordinate =coordinate).value
            except Exception as e:
                raise
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def getCellOfObject(self,sheet,coordinate = None,rowNo = None,colsNo = None):
        if coordinate!=None:
            try:
                return sheet.cell(coordinate =coordinate)
            except Exception as e:
                raise
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def writeCell(self,sheet,content,coordinate = None,rowNo = None,colsNo = None,style = None):
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content  #将content内容赋值给sheet.cell.value
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style]) #给单元格上色
                self.workbook.save(self.excelFile) #保存输入的content和颜色
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = content  # 将content内容赋值给sheet.cell.value
                if style is not None:
                    sheet.cell(row=rowNo,column=colsNo).font = Font(color=self.RGBDict[style])  # 给单元格上色
                self.workbook.save(self.excelFile)  # 保存输入的content和颜色
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        now =int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime('%Y-%m-%d %H:%M:%S',timeArray )
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime  # 将currentTime赋值给sheet.cell.value
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value = currentTime  # 将currentTime赋值给sheet.cell.value
                self.workbook.save(self.excelFile)  # 保存输入
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell")

if __name__== '__main__':
    pe = PaseExcel()
    pe.loadWorkBook('C:\\test\\testexcel.xlsx')

    #print("通过name获取sheet对象名:",pe.getSheetByName("联系人").title)
    #print("通过index获取sheet对象名:",pe.getSheetByIndex(0).title)
    sheet =pe.getSheetByIndex(0)
    #print(type(sheet))
    #print(pe.getRowsNumber(sheet))
    #print(pe.getColsNumber(sheet))
    #rows = pe.getRow(sheet,1)
    #for i in rows:
    #   print(i.value)
   # print(pe.getCellOfValue(sheet,rowNo=1,colsNo=1))

    pe.writeCell(sheet,'我是一条测试文本',rowNo=10,colsNo=10,style='green')
    pe.writeCellCurrentTime(sheet,rowNo=10,colsNo=11)